import os
import pandas as pd
from flask import Flask, request, render_template, send_file
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill

app = Flask(__name__)

# --- CONFIGURATION ---
# In production (Render), these files must be inside the repo.
DATABASE_FOLDER = 'database_files'

# Group 1: IMNEO (Always RED)
IMNEO_FILES = [
    'imneo.csv',
    'imneo1.xlsx',
    'imneo2.xlsx'
]

# Group 2: X-CLIENTS (Yellow or Red depending on mode)
XCLIENT_FILES = [
    'xclient1.csv',
    'xclient2.csv'
]

def normalize_text(text):
    """Cleans text for smart matching (lowercase, stripped)."""
    if pd.isna(text) or text is None or str(text).lower() == 'nan':
        return ""
    return str(text).lower().strip()

def load_database_set(filenames):
    """Loads multiple CSV/Excel files into a single set of names and companies."""
    names_set = set()
    companies_set = set()

    if not os.path.exists(DATABASE_FOLDER):
        print(f"CRITICAL ERROR: Folder '{DATABASE_FOLDER}' not found.")
        return names_set, companies_set

    for filename in filenames:
        path = os.path.join(DATABASE_FOLDER, filename)
        if not os.path.exists(path):
            print(f"Warning: Database file not found: {path}")
            continue

        try:
            if filename.endswith('.csv'):
                try:
                    df = pd.read_csv(path, encoding='utf-8', on_bad_lines='skip', sep=None, engine='python')
                except:
                    df = pd.read_csv(path, encoding='latin1', on_bad_lines='skip', sep=None, engine='python')
            else:
                df = pd.read_excel(path)
            
            df.columns = [str(c).lower().strip() for c in df.columns]

            for _, row in df.iterrows():
                first, last, company = "", "", ""

                # Header Mapping
                if 'first name' in row: first = row['first name']
                elif 'voor naam' in row: first = row['voor naam']
                elif 'naam' in row: first = row['naam']

                if 'last name' in row: last = row['last name']
                elif 'achternaam' in row: last = row['achternaam']

                if 'company name' in row: company = row['company name']
                elif 'company' in row: company = row['company']
                elif 'huidig bedrijf' in row: company = row['huidig bedrijf']
                elif 'currrent company' in row: company = row['currrent company']
                elif 'current company' in row: company = row['current company']
                elif 'company table data' in row: company = row['company table data']

                full_name = f"{normalize_text(first)} {normalize_text(last)}".strip()
                if full_name: names_set.add(full_name)
                
                norm_company = normalize_text(company)
                if norm_company: companies_set.add(norm_company)

        except Exception as e:
            print(f"Error loading {filename}: {e}")

    return names_set, companies_set

# --- LOAD DATABASES ON STARTUP ---
# Only run this if we are not in a build step (prevents errors during deployment build)
if os.environ.get('RENDER') is None or os.path.exists(DATABASE_FOLDER):
    print("Loading Databases...")
    imneo_names, imneo_companies = load_database_set(IMNEO_FILES)
    xclient_names, xclient_companies = load_database_set(XCLIENT_FILES)
else:
    print("Skipping database load during build.")
    imneo_names, imneo_companies = set(), set()
    xclient_names, xclient_companies = set(), set()

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files: return "No file uploaded", 400
    file = request.files['file']
    mode = request.form.get('mode')

    if file.filename == '': return "No file selected", 400

    try:
        if file.filename.endswith('.csv'):
            try: df = pd.read_csv(file, encoding='utf-8', sep=None, engine='python')
            except: 
                file.seek(0)
                df = pd.read_csv(file, encoding='latin1', sep=None, engine='python')
        else:
            df = pd.read_excel(file)
    except Exception as e:
        return f"Error reading file: {e}", 500

    # Column Detection
    first_name_col = next((c for c in df.columns if str(c).lower().strip() in ['first name', 'voor naam', 'naam']), None)
    last_name_col = next((c for c in df.columns if str(c).lower().strip() in ['last name', 'achternaam']), None)
    
    company_col = None
    possible_company_headers = ['company table data', 'company name', 'huidig bedrijf', 'company', 'name']
    for header in possible_company_headers:
        match = next((c for c in df.columns if str(c).lower().strip() == header), None)
        if match:
            company_col = match
            break

    results = []
    colors = []

    for index, row in df.iterrows():
        user_name = ""
        if first_name_col and last_name_col:
             user_name = f"{normalize_text(row[first_name_col])} {normalize_text(row[last_name_col])}".strip()
        
        user_company = normalize_text(row[company_col]) if company_col else ""
        
        status = "Safe"
        color_code = "FFFFFF"

        is_imneo = (user_name and user_name in imneo_names) or (user_company and user_company in imneo_companies)
        is_xclient = (user_name and user_name in xclient_names) or (user_company and user_company in xclient_companies)

        if is_imneo:
            status = "IMNEO Match (Restricted)"
            color_code = "FF0000"
        elif is_xclient:
            if mode == 'candidate':
                status = "X-Client Match (Candidate Mode)"
                color_code = "FF0000"
            else:
                status = "X-Client Match (Client/Relation)"
                color_code = "FFFF00"
        
        results.append(status)
        colors.append(color_code)

    df['Check Result'] = results
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        for i, color in enumerate(colors):
            if color != "FFFFFF":
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for col in range(1, len(df.columns) + 1):
                    worksheet.cell(row=i+2, column=col).fill = fill

    output.seek(0)
    return send_file(output, download_name="checked_list.xlsx", as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)